"""
導出服務 - 處理 US-005
支持 PDF、Excel、CSV 格式導出
"""

import io
import csv
import zipfile
import matplotlib.pyplot as plt
from datetime import datetime
from pathlib import Path
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from gradeinsight.models import Student, Grade, GradeComponent, ReportTemplate


class ExportService:
    """導出服務"""
    
    @staticmethod
    def export_class_report_excel(output_path: str) -> str:
        """
        導出班級成績表 - Excel 格式
        
        Returns: 文件路徑
        """
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "班級成績表"
        
        # 設置標題
        title_cell = worksheet['A1']
        title_cell.value = "班級成績表"
        title_cell.font = Font(size=16, bold=True)
        worksheet.merge_cells('A1:G1')
        
        worksheet['A2'] = f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 表頭
        headers = ['學號', '姓名', '班級']
        components = GradeComponent.query.all()
        for comp in components:
            headers.append(comp.name)
        headers.append('最終分數')
        
        worksheet.append(headers)
        
        # 設置表頭樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # 數據行
        students = Student.query.all()
        for student in students:
            row = [student.student_id, student.name, student.class_name or '']
            
            total_weighted = 0
            total_weight = 0
            
            for component in components:
                grade = Grade.query.filter_by(
                    student_id=student.id,
                    component_id=component.id
                ).first()
                
                if grade:
                    row.append(grade.score)
                    total_weighted += grade.score * component.weight
                    total_weight += component.weight
                else:
                    row.append('-')
            
            final_score = total_weighted / total_weight if total_weight > 0 else 0
            row.append(round(final_score, 2))
            
            worksheet.append(row)
        
        # 調整列寬
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 20)
        
        workbook.save(output_path)
        return output_path
    
    @staticmethod
    def export_individual_report_excel(student_id: int, output_path: str) -> str:
        """導出個人成績單 - Excel 格式"""
        
        student = Student.query.get(student_id)
        if not student:
            raise ValueError(f"學生 ID {student_id} 不存在")
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "成績單"
        
        # 標題
        title = worksheet['A1']
        title.value = "個人成績單"
        title.font = Font(size=18, bold=True)
        worksheet.merge_cells('A1:C1')
        
        # 學生信息
        worksheet['A3'] = "學號:"
        worksheet['B3'] = student.student_id
        worksheet['A4'] = "姓名:"
        worksheet['B4'] = student.name
        worksheet['A5'] = "班級:"
        worksheet['B5'] = student.class_name or '-'
        worksheet['A6'] = "生成日期:"
        worksheet['B6'] = datetime.now().strftime('%Y-%m-%d')
        
        # 成績明細
        worksheet['A8'] = "成績明細"
        worksheet['A8'].font = Font(size=12, bold=True)
        
        headers = ['組件', '分數', '權重', '加權分']
        worksheet.append([])
        worksheet.append(headers)
        
        # 表頭樣式
        header_row = worksheet[10]
        for cell in header_row:
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)
        
        # 數據
        total_weighted = 0
        total_weight = 0
        row_num = 11
        
        components = GradeComponent.query.all()
        for component in components:
            grade = Grade.query.filter_by(
                student_id=student_id,
                component_id=component.id
            ).first()
            
            if grade:
                worksheet[f'A{row_num}'] = component.name
                worksheet[f'B{row_num}'] = grade.score
                worksheet[f'C{row_num}'] = component.weight
                worksheet[f'D{row_num}'] = round(grade.score * component.weight, 2)
                
                total_weighted += grade.score * component.weight
                total_weight += component.weight
                row_num += 1
        
        # 最終分數
        final_score = total_weighted / total_weight if total_weight > 0 else 0
        worksheet[f'A{row_num+1}'] = "最終分數:"
        worksheet[f'A{row_num+1}'].font = Font(bold=True)
        worksheet[f'B{row_num+1}'] = round(final_score, 2)
        worksheet[f'B{row_num+1}'].font = Font(bold=True)
        
        # 調整列寬
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 12
        worksheet.column_dimensions['D'].width = 15
        
        workbook.save(output_path)
        return output_path
    
    @staticmethod
    def export_class_report_csv(output_path: str) -> str:
        """導出班級成績表 - CSV 格式"""
        
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # 表頭
            headers = ['學號', '姓名', '班級']
            components = GradeComponent.query.all()
            for comp in components:
                headers.append(comp.name)
            headers.append('最終分數')
            
            writer.writerow(headers)
            
            # 數據
            students = Student.query.all()
            for student in students:
                row = [student.student_id, student.name, student.class_name or '']
                
                total_weighted = 0
                total_weight = 0
                
                for component in components:
                    grade = Grade.query.filter_by(
                        student_id=student.id,
                        component_id=component.id
                    ).first()
                    
                    if grade:
                        row.append(grade.score)
                        total_weighted += grade.score * component.weight
                        total_weight += component.weight
                    else:
                        row.append('-')
                
                final_score = total_weighted / total_weight if total_weight > 0 else 0
                row.append(round(final_score, 2))
                
                writer.writerow(row)
        
        return output_path
    
    @staticmethod
    def export_statistics_summary(output_path: str) -> str:
        """導出統計摘要 - Excel 格式"""
        
        from gradeinsight.services.analysis import GradeAnalysisService
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "統計分析"
        
        # 標題
        title = worksheet['A1']
        title.value = "成績統計分析報告"
        title.font = Font(size=16, bold=True)
        worksheet.merge_cells('A1:C1')
        
        worksheet['A2'] = f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 基本統計
        worksheet['A4'] = "基本統計指標"
        worksheet['A4'].font = Font(size=12, bold=True)
        
        summary = GradeAnalysisService.get_statistics_summary()
        
        row_num = 5
        metrics = [
            ('總學生數', summary['total_students']),
            ('平均分', summary['statistics'].get('mean', 0)),
            ('中位數', summary['statistics'].get('median', 0)),
            ('標準差', summary['statistics'].get('std_dev', 0)),
            ('最高分', summary['statistics'].get('max', 0)),
            ('最低分', summary['statistics'].get('min', 0)),
            ('及格率 (%)', summary['pass_rate']),
            ('優秀率 (%)', summary['excellent_rate']),
        ]
        
        for metric, value in metrics:
            worksheet[f'A{row_num}'] = metric
            worksheet[f'B{row_num}'] = value
            row_num += 1
        
        # 成績分級
        worksheet[f'A{row_num+1}'] = "成績分級分布"
        worksheet[f'A{row_num+1}'].font = Font(size=12, bold=True)
        
        row_num += 2
        worksheet[f'A{row_num}'] = "等級"
        worksheet[f'B{row_num}'] = "人數"
        row_num += 1
        
        for level, count in summary['grade_levels'].items():
            worksheet[f'A{row_num}'] = level
            worksheet[f'B{row_num}'] = count
            row_num += 1
        
        # 調整列寬
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        
        workbook.save(output_path)
        return output_path

    @staticmethod
    def export_class_report_pdf(output_path: str) -> str:
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('班級成績報告', styles['Title']))
        elements.append(Spacer(1, 12))

        headers = ['學號', '姓名', '班級']
        components = GradeComponent.query.all()
        for comp in components:
            headers.append(comp.name)
        headers.append('最終分數')

        data = [headers]
        students = Student.query.all()
        for student in students:
            row = [student.student_id, student.name, student.class_name or '']
            total_weighted = 0
            total_weight = 0
            for component in components:
                grade = Grade.query.filter_by(student_id=student.id, component_id=component.id).first()
                if grade:
                    row.append(round(grade.score, 2))
                    total_weighted += grade.score * component.weight
                    total_weight += component.weight
                else:
                    row.append('-')
            final_score = total_weighted / total_weight if total_weight > 0 else 0
            row.append(round(final_score, 2))
            data.append(row)

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        doc.build(elements)
        return output_path

    @staticmethod
    def export_individual_report_pdf(student_id: int, output_path: str) -> str:
        student = Student.query.get(student_id)
        if not student:
            raise ValueError(f"學生 ID {student_id} 不存在")

        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('個人成績單', styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f'學號: {student.student_id}', styles['Normal']))
        elements.append(Paragraph(f'姓名: {student.name}', styles['Normal']))
        elements.append(Paragraph(f'班級: {student.class_name or "-"}', styles['Normal']))
        elements.append(Paragraph(f'生成日期: {datetime.now().strftime("%Y-%m-%d")}', styles['Normal']))
        elements.append(Spacer(1, 12))

        table_data = [['組件', '分數', '權重', '加權分']]
        total_weighted = 0
        total_weight = 0
        components = GradeComponent.query.all()
        for component in components:
            grade = Grade.query.filter_by(student_id=student_id, component_id=component.id).first()
            if grade:
                weighted_score = grade.score * component.weight
                table_data.append([
                    component.name,
                    round(grade.score, 2),
                    round(component.weight, 2),
                    round(weighted_score, 2),
                ])
                total_weighted += weighted_score
                total_weight += component.weight

        final_score = total_weighted / total_weight if total_weight > 0 else 0
        table_data.append(['', '', '最終分數', round(final_score, 2)])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        doc.build(elements)
        return output_path

    @staticmethod
    def export_batch_individual_reports_zip(output_path: str) -> str:
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for student in Student.query.all():
                pdf_name = f'{student.student_id}_{student.name}.pdf'
                pdf_path = Path(output_path).parent / pdf_name
                ExportService.export_individual_report_pdf(student.id, str(pdf_path))
                zf.write(str(pdf_path), pdf_name)
                pdf_path.unlink()
        return output_path

    @staticmethod
    def export_grade_distribution_png(output_path: str, bin_width: int = 10, component_id: int = None) -> str:
        from gradeinsight.services.analysis import GradeAnalysisService
        dist = GradeAnalysisService.get_grade_distribution(component_id=component_id, bin_width=bin_width)
        if not dist['bins']:
            raise ValueError('沒有成績資料可產生圖表')

        plt.figure(figsize=(8, 4))
        counts = dist['counts']
        plt.bar(dist['bins'], counts, color='skyblue')
        plt.xlabel('成績段位')
        plt.ylabel('人數')
        plt.title('成績分佈直方圖')
        plt.tight_layout()
        plt.savefig(output_path)
        plt.close()
        return output_path

    @staticmethod
    def export_statistics_pdf(output_path: str) -> str:
        from gradeinsight.services.analysis import GradeAnalysisService
        summary = GradeAnalysisService.get_statistics_summary()
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('成績統計分析報告', styles['Title']))
        elements.append(Spacer(1, 12))
        table_data = [['指標', '數值']]
        for key, value in summary['statistics'].items():
            table_data.append([key, value])
        table_data.append(['及格率 (%)', summary['pass_rate']])
        table_data.append(['優良率 (%)', summary['excellent_rate']])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        doc.build(elements)
        return output_path

    @staticmethod
    def export_report_with_template(template_id: int, output_path: str) -> str:
        template = ReportTemplate.query.get(template_id)
        if not template:
            raise ValueError(f'報表範本 {template_id} 不存在')

        fields = template.get_fields()
        rows = []
        headers = fields
        for student in Student.query.all():
            row = []
            for field in fields:
                if field == '學號':
                    row.append(student.student_id)
                elif field == '姓名':
                    row.append(student.name)
                elif field == '班級':
                    row.append(student.class_name or '')
                elif field == '最終成績':
                    grades = Grade.query.filter_by(student_id=student.id).all()
                    total_weighted = 0
                    total_weight = 0
                    for grade in grades:
                        comp = GradeComponent.query.get(grade.component_id)
                        if comp:
                            total_weighted += grade.score * comp.weight
                            total_weight += comp.weight
                    final_score = total_weighted / total_weight if total_weight > 0 else 0
                    row.append(round(final_score, 2))
                elif '各科成績' in field:
                    component_scores = []
                    for grade in Grade.query.filter_by(student_id=student.id).all():
                        comp = GradeComponent.query.get(grade.component_id)
                        if comp:
                            component_scores.append(f"{comp.name}:{round(grade.score,2)}")
                    row.append('; '.join(component_scores))
                elif '條碼' in field:
                    row.append(f"QR:{student.student_id}")
                elif '評語' in field:
                    row.append('請持續保持學習動力')
                else:
                    row.append('')
            rows.append(row)

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            writer.writerows(rows)

        return output_path
